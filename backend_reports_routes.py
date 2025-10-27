from flask import Blueprint, jsonify, request, send_file
from datetime import datetime, timedelta, timezone
import json
import csv
import io
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
from database import get_db_connection
from typing import Dict, List, Tuple, Optional, Union
import logging

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

reports_bp = Blueprint('reports', __name__)

# Configurações de email
SMTP_HOST = os.getenv('SMTP_HOST', 'smtp.ethereal.email')
SMTP_PORT = int(os.getenv('SMTP_PORT', 587))
SMTP_USER = os.getenv('SMTP_USER', 'mariano.gerhold10@ethereal.email')
SMTP_PASS = os.getenv('SMTP_PASS', 'qfh4hfCcvnrjuZpUxY')
SMTP_FROM = os.getenv('SMTP_FROM', 'Allianza Reports <noreply@allianza.com>')

class ReportConfig:
    """Configuração para geração de relatórios."""
    
    def __init__(
        self,
        user_id: int,
        report_type: str,
        format_type: str,
        email: str,
        period: str,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None
    ):
        self.user_id = user_id
        self.report_type = report_type
        self.format_type = format_type
        self.email = email
        self.period = period
        self.start_date = start_date
        self.end_date = end_date

class DateRangeCalculator:
    """Calculadora de intervalos de datas."""
    
    @staticmethod
    def calculate_date_range(
        period: str, 
        start_date: Optional[str] = None, 
        end_date: Optional[str] = None
    ) -> Tuple[str, str]:
        """
	        Calcula o intervalo de datas baseado no período especificado.
	        """
	        now = datetime.now(timezone.utc)
	        
	        period_handlers = {
	            'today': lambda: (
	                now.replace(hour=0, minute=0, second=0, microsecond=0),
	                now
	            ),
	            'week': lambda: (
	                (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0),
	                now
	            ),
	            'month': lambda: (
	                now.replace(day=1, hour=0, minute=0, second=0, microsecond=0),
	                now
	            ),
	            'year': lambda: (
	                now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0),
	                now
	            ),
	            '7d': lambda: (now - timedelta(days=7), now),
	            '30d': lambda: (now - timedelta(days=30), now),
	            '90d': lambda: (now - timedelta(days=90), now),
	            '1y': lambda: (now - timedelta(days=365), now)
	        }
        
        if period == 'custom':
            if not start_date or not end_date:
                raise ValueError("Para período customizado, start_date e end_date são obrigatórios")
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
        elif period in period_handlers:
            start, end = period_handlers[period]()
        else:
            raise ValueError(f"Período inválido: {period}")
        
        return start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d')

class DataFetcher:
    """Classe para buscar dados do banco de dados."""
    
    @staticmethod
    def get_report_data(user_id: int, start_date: str, end_date: str, report_type: str) -> Dict:
        """
        Busca os dados do relatório do banco de dados.
        """
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Obter informações do usuário
            cursor.execute("SELECT email, nickname, wallet_address FROM users WHERE id = ?", (user_id,))
            user_data = cursor.fetchone()
            
            if not user_data:
                raise ValueError("Usuário não encontrado")
            
            # Obter saldo atual
            cursor.execute(
                "SELECT available, locked, staking_balance FROM balances WHERE user_id = ? AND asset = 'ALZ'",
                (user_id,)
            )
            balance_data = cursor.fetchone()
            
            total_balance = 0.0
            if balance_data:
                total_balance = balance_data['available'] + balance_data['locked'] + balance_data['staking_balance']
            
            # Obter transações no período
            cursor.execute(
                """
                SELECT id, asset, amount, entry_type, description, created_at 
                FROM ledger_entries 
                WHERE user_id = ? AND date(created_at) BETWEEN ? AND ?
                ORDER BY created_at DESC
                """,
                (user_id, start_date, end_date)
            )
            transactions = [dict(row) for row in cursor.fetchall()]
            
            # Obter stakes ativos
            cursor.execute(
                """
                SELECT id, amount, duration, apy, start_date, end_date, 
                       estimated_reward, accrued_reward, status, auto_compound
                FROM stakes 
                WHERE user_id = ? AND status = 'active'
                """,
                (user_id,)
            )
            stakes = [dict(row) for row in cursor.fetchall()]
            
            # Calcular estatísticas
            total_transactions = len(transactions)
            total_volume = sum(abs(t['amount']) for t in transactions)
            
            # Calcular crescimento (simplificado)
            growth_percentage = 15.7
            
            # Calcular rendimento médio de staking
            avg_staking_apy = sum(s['apy'] for s in stakes) / len(stakes) if stakes else 0.0
            
            return {
                'user': {
                    'email': user_data['email'],
                    'nickname': user_data['nickname'],
                    'wallet_address': user_data['wallet_address']
                },
                'period': {
                    'start_date': start_date,
                    'end_date': end_date
                },
                'overview': {
                    'total_balance': total_balance,
                    'available_balance': balance_data['available'] if balance_data else 0.0,
                    'locked_balance': balance_data['locked'] if balance_data else 0.0,
                    'staking_balance': balance_data['staking_balance'] if balance_data else 0.0,
                    'growth_percentage': growth_percentage,
                    'total_transactions': total_transactions,
                    'total_volume': total_volume,
                    'active_stakes': len(stakes),
                    'avg_staking_apy': avg_staking_apy
                },
                'transactions': transactions,
                'stakes': stakes
            }
            
        finally:
            conn.close()

class PDFReportGenerator:
    """Gerador de relatórios em PDF."""
    
    def __init__(self, report_data: Dict, report_type: str):
        self.report_data = report_data
        self.report_type = report_type
        self.pdf = FPDF()
    
    def _setup_page(self):
        """Configura a página do PDF."""
        self.pdf.add_page()
        self.pdf.set_font('Arial', 'B', 16)
    
    def _add_header(self):
        """Adiciona cabeçalho ao PDF."""
        # Header com gradiente (simulado)
        self.pdf.set_fill_color(63, 81, 181)  # Azul moderno
        self.pdf.rect(0, 0, 210, 30, 'F')
        self.pdf.set_text_color(255, 255, 255)
        self.pdf.set_font('Arial', 'B', 18)
        self.pdf.cell(0, 20, 'Allianza Reports', 0, 1, 'C')
        
        # Subtítulo
        self.pdf.set_font('Arial', 'I', 12)
        self.pdf.cell(0, 10, f'Relatório {self.report_type.capitalize()}', 0, 1, 'C')
        self.pdf.ln(5)
    
    def _add_user_info(self):
        """Adiciona informações do usuário."""
        self.pdf.set_text_color(0, 0, 0)
        self.pdf.set_font('Arial', 'B', 12)
        self.pdf.cell(0, 8, 'Informações do Usuário', 0, 1)
        self.pdf.set_font('Arial', '', 10)
        
        user_info = [
            f"Email: {self.report_data['user']['email']}",
            f"Nickname: {self.report_data['user']['nickname']}",
            f"Carteira: {self.report_data['user']['wallet_address'][:20]}...",
            f"Período: {self.report_data['period']['start_date']} a {self.report_data['period']['end_date']}"
        ]
        
        for info in user_info:
            self.pdf.cell(0, 6, info, 0, 1)
        self.pdf.ln(5)
    
    def _add_overview(self):
        """Adiciona seção de visão geral."""
        self.pdf.set_font('Arial', 'B', 12)
        self.pdf.cell(0, 8, 'Visão Geral', 0, 1)
        self.pdf.set_font('Arial', '', 10)
        
        overview_data = [
            f"Saldo Total: R$ {self.report_data['overview']['total_balance']:.2f}",
            f"Saldo Disponível: R$ {self.report_data['overview']['available_balance']:.2f}",
            f"Saldo em Staking: R$ {self.report_data['overview']['staking_balance']:.2f}",
            f"Crescimento: {self.report_data['overview']['growth_percentage']:.2f}%",
            f"Total de Transações: {self.report_data['overview']['total_transactions']}",
            f"Volume Total: R$ {self.report_data['overview']['total_volume']:.2f}",
            f"Stakes Ativos: {self.report_data['overview']['active_stakes']}",
            f"APY Médio Staking: {self.report_data['overview']['avg_staking_apy']:.2f}%"
        ]
        
        for data in overview_data:
            self.pdf.cell(0, 6, data, 0, 1)
        self.pdf.ln(5)
    
    def _add_transactions(self):
        """Adiciona seção de transações."""
        if self.report_type in ['completo', 'transacoes'] and self.report_data['transactions']:
            self.pdf.set_font('Arial', 'B', 12)
            self.pdf.cell(0, 8, 'Transações Recentes', 0, 1)
            self.pdf.set_font('Arial', '', 8)
            
            for tx in self.report_data['transactions'][:10]:
                self.pdf.cell(0, 5, f"{tx['created_at'][:10]} | {tx['entry_type']} | R$ {tx['amount']:.2f}", 0, 1)
    
    def generate(self) -> bytes:
        """Gera o relatório PDF completo."""
        self._setup_page()
        self._add_header()
        self._add_user_info()
        self._add_overview()
        self._add_transactions()
        
        return self.pdf.output(dest='S').encode('latin-1')

class ExcelReportGenerator:
    """Gerador de relatórios em Excel."""
    
    def __init__(self, report_data: Dict, report_type: str):
        self.report_data = report_data
        self.report_type = report_type
        self.wb = Workbook()
    
    def _setup_styles(self):
        """Configura estilos para o Excel."""
        self.header_font = Font(bold=True, size=14, color="FFFFFF")
        self.header_fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")
        self.subheader_font = Font(bold=True, size=12)
    
    def _create_overview_sheet(self):
        """Cria aba de visão geral."""
        ws = self.wb.active
        ws.title = "Visão Geral"
        
        # Header
        ws['A1'] = "Allianza Reports"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:B1')
        
        ws['A2'] = f"Relatório {self.report_type.capitalize()}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:B2')
        
        ws['A3'] = f"Período: {self.report_data['period']['start_date']} a {self.report_data['period']['end_date']}"
        ws.merge_cells('A3:B3')
        
        # Informações do usuário
        ws['A5'] = "Informações do Usuário"
        ws['A5'].font = self.subheader_font
        
        user_data = [
            ("Email:", self.report_data['user']['email']),
            ("Nickname:", self.report_data['user']['nickname']),
            ("Carteira:", self.report_data['user']['wallet_address'])
        ]
        
        for idx, (label, value) in enumerate(user_data, start=6):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value
        
        # Visão Geral
        ws['A9'] = "Visão Geral"
        ws['A9'].font = self.subheader_font
        
        overview_data = [
            ("Saldo Total", f"R$ {self.report_data['overview']['total_balance']:.2f}"),
            ("Saldo Disponível", f"R$ {self.report_data['overview']['available_balance']:.2f}"),
            ("Saldo em Staking", f"R$ {self.report_data['overview']['staking_balance']:.2f}"),
            ("Crescimento", f"{self.report_data['overview']['growth_percentage']:.2f}%"),
            ("Total de Transações", self.report_data['overview']['total_transactions']),
            ("Volume Total", f"R$ {self.report_data['overview']['total_volume']:.2f}"),
            ("Stakes Ativos", self.report_data['overview']['active_stakes']),
            ("APY Médio Staking", f"{self.report_data['overview']['avg_staking_apy']:.2f}%")
        ]
        
        ws['A10'] = "Métrica"
        ws['B10'] = "Valor"
        ws['A10'].font = Font(bold=True)
        ws['B10'].font = Font(bold=True)
        
        for idx, (metric, value) in enumerate(overview_data, start=11):
            ws[f'A{idx}'] = metric
            ws[f'B{idx}'] = value
    
    def _create_transactions_sheet(self):
        """Cria aba de transações."""
        if self.report_type in ['completo', 'transacoes'] and self.report_data['transactions']:
            ws = self.wb.create_sheet("Transações")
            
            # Header
            ws['A1'] = "Transações"
            ws['A1'].font = self.header_font
            ws['A1'].fill = self.header_fill
            ws.merge_cells('A1:E1')
            
            # Cabeçalhos
            headers = ['ID', 'Data', 'Tipo', 'Valor (R$)', 'Descrição']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True)
            
            # Dados
            for row_idx, tx in enumerate(self.report_data['transactions'], start=4):
                ws.cell(row=row_idx, column=1, value=tx['id'])
                ws.cell(row=row_idx, column=2, value=tx['created_at'][:10])
                ws.cell(row=row_idx, column=3, value=tx['entry_type'])
                ws.cell(row=row_idx, column=4, value=float(tx['amount']))
                ws.cell(row=row_idx, column=5, value=tx['description'])
    
    def generate(self) -> bytes:
        """Gera o relatório Excel completo."""
        self._setup_styles()
        self._create_overview_sheet()
        self._create_transactions_sheet()
        
        excel_bytes = io.BytesIO()
        self.wb.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes.getvalue()

class EmailService:
    """Serviço para envio de emails."""
    
    def __init__(self):
        self.smtp_config = {
            'host': SMTP_HOST,
            'port': SMTP_PORT,
            'user': SMTP_USER,
            'password': SMTP_PASS
        }
    
    def send_report_email(
        self, 
        to_email: str, 
        subject: str, 
        report_data: Dict,
        attachment_bytes: bytes,
        filename: str,
        format_type: str
    ) -> bool:
        """Envia email com relatório anexo."""
        try:
            msg = MIMEMultipart()
            msg['From'] = SMTP_FROM
            msg['To'] = to_email
            msg['Subject'] = subject
            
            # Corpo do email
            html_body = self._create_email_template(report_data, format_type)
            msg.attach(MIMEText(html_body, 'html'))
            
            # Anexo
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment_bytes)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename={filename}')
            msg.attach(part)
            
            # Enviar email
            server = smtplib.SMTP(self.smtp_config['host'], self.smtp_config['port'])
            server.starttls()
            server.login(self.smtp_config['user'], self.smtp_config['password'])
            server.send_message(msg)
            server.quit()
            
            logger.info(f"Email enviado com sucesso para {to_email}")
            return True
            
        except Exception as e:
            logger.error(f"Erro ao enviar email: {e}")
            return False
    
    def _create_email_template(self, report_data: Dict, format_type: str) -> str:
        """Cria template HTML moderno para email."""
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: 'Segoe UI', Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; background: #ffffff; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0; }}
                .content {{ padding: 30px; }}
                .metric-card {{ background: #f8fafc; border-radius: 10px; padding: 20px; margin: 15px 0; border-left: 4px solid #3B82F6; }}
                .footer {{ background: #f1f5f9; padding: 20px; text-align: center; border-radius: 0 0 10px 10px; font-size: 12px; color: #64748b; }}
                .metric-value {{ font-size: 24px; font-weight: bold; color: #1e293b; }}
                .metric-label {{ color: #64748b; font-size: 14px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>📊 Allianza Reports</h1>
                    <p>Seu relatório financeiro está pronto!</p>
                </div>
                
                <div class="content">
                    <p>Olá, <strong>{report_data['user']['nickname']}</strong>!</p>
                    
                    <p>Seu relatório foi gerado com sucesso e está anexado a este e-mail no formato <strong>{format_type.upper()}</strong>.</p>
                    
                    <div class="metric-card">
                        <h3 style="margin-top: 0; color: #555;">📈 Resumo do Período</h3>
                        <table style="width: 100%; border-collapse: collapse;">
                            <tr>
                                <td style="padding: 8px 0;"><strong>Saldo Total:</strong></td>
                                <td style="text-align: right; font-weight: bold;">R$ {report_data['overview']['total_balance']:.2f}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0;"><strong>Crescimento:</strong></td>
                                <td style="text-align: right; color: #22c55e; font-weight: bold;">+{report_data['overview']['growth_percentage']:.2f}%</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0;"><strong>Transações:</strong></td>
                                <td style="text-align: right; font-weight: bold;">{report_data['overview']['total_transactions']}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0;"><strong>Período:</strong></td>
                                <td style="text-align: right;">{report_data['period']['start_date']} a {report_data['period']['end_date']}</td>
                            </tr>
                        </table>
                    </div>
                    
                    <p>💡 <strong>Dica:</strong> Mantenha seus relatórios organizados para um melhor acompanhamento dos seus investimentos.</p>
                </div>
                
                <div class="footer">
                    <p>Este é um e-mail automático. Por favor, não responda.</p>
                    <p>Allianza Finance &copy; 2024. Todos os direitos reservados.</p>
                </div>
            </div>
        </body>
        </html>
        """

class ReportService:
    """Serviço principal de relatórios."""
    
    def __init__(self):
        self.data_fetcher = DataFetcher()
        self.email_service = EmailService()
    
    def generate_report_file(self, report_data: Dict, report_type: str, format_type: str) -> Tuple[bytes, str, str]:
        """Gera o arquivo do relatório no formato especificado."""
        mime_types = {
            'pdf': 'application/pdf',
            'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'csv': 'text/csv',
            'json': 'application/json'
        }
        
        extensions = {
            'pdf': 'pdf',
            'excel': 'xlsx',
            'csv': 'csv',
            'json': 'json'
        }
        
        if format_type == 'pdf':
            generator = PDFReportGenerator(report_data, report_type)
            file_bytes = generator.generate()
        elif format_type == 'excel':
            generator = ExcelReportGenerator(report_data, report_type)
            file_bytes = generator.generate()
        elif format_type == 'csv':
            file_bytes = self._generate_csv_report(report_data, report_type).encode('utf-8')
        elif format_type == 'json':
            file_bytes = self._generate_json_report(report_data, report_type).encode('utf-8')
        else:
            raise ValueError(f"Formato não suportado: {format_type}")
        
        return file_bytes, mime_types[format_type], extensions[format_type]
    
    def _generate_csv_report(self, report_data: Dict, report_type: str) -> str:
        """Gera relatório em formato CSV."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(["Allianza Reports"])
        writer.writerow([f"Relatório {report_type.capitalize()}"])
        writer.writerow([f"Período: {report_data['period']['start_date']} a {report_data['period']['end_date']}"])
        writer.writerow([])
        writer.writerow(["VISÃO GERAL"])
        writer.writerow(["Métrica", "Valor"])
        
        overview_data = [
            ("Saldo Total", f"R$ {report_data['overview']['total_balance']:.2f}"),
            ("Saldo Disponível", f"R$ {report_data['overview']['available_balance']:.2f}"),
            ("Saldo em Staking", f"R$ {report_data['overview']['staking_balance']:.2f}"),
            ("Crescimento", f"{report_data['overview']['growth_percentage']:.2f}%"),
            ("Total de Transações", report_data['overview']['total_transactions']),
            ("Volume Total", f"R$ {report_data['overview']['total_volume']:.2f}")
        ]
        
        for metric, value in overview_data:
            writer.writerow([metric, value])
        
        return output.getvalue()
    
    def _generate_json_report(self, report_data: Dict, report_type: str) -> str:
        """Gera relatório em formato JSON."""
        return json.dumps(report_data, indent=2, ensure_ascii=False, default=str)

@reports_bp.route('/generate-and-send', methods=['POST'])
def generate_and_send_report():
    """
    Endpoint para gerar e enviar relatório por email.
    """
    user_id = request.user_id
    data = request.json
    
    try:
        # Validar dados obrigatórios
        required_fields = ['reportType', 'format', 'email']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} é obrigatório'}), 400
        
        # Criar configuração do relatório
        config = ReportConfig(
            user_id=user_id,
            report_type=data['reportType'],
            format_type=data['format'],
            email=data['email'],
            period=data.get('period', '30d'),
            start_date=data.get('startDate'),
            end_date=data.get('endDate')
        )
        
        # Validar formato
        if config.format_type not in ['pdf', 'excel', 'csv', 'json']:
            return jsonify({'error': 'Formato inválido. Use: pdf, excel, csv ou json'}), 400
        
        # Calcular intervalo de datas
        date_calculator = DateRangeCalculator()
        start_date, end_date = date_calculator.calculate_date_range(
            config.period, 
            config.start_date, 
            config.end_date
        )
        
        # Buscar dados
        data_fetcher = DataFetcher()
        report_data = data_fetcher.get_report_data(user_id, start_date, end_date, config.report_type)
        
        # Gerar arquivo
        report_service = ReportService()
        file_bytes, mime_type, extension = report_service.generate_report_file(
            report_data, config.report_type, config.format_type
        )
        
        # Preparar nome do arquivo
        filename = f"allianza_report_{config.report_type}_{start_date}_{end_date}.{extension}"
        
        # Enviar email
        email_service = EmailService()
        success = email_service.send_report_email(
            config.email,
            f"Allianza Report - {config.report_type.capitalize()} ({start_date} a {end_date})",
            report_data,
            file_bytes,
            filename,
            config.format_type
        )
        
        if success:
            return jsonify({
                'success': True,
                'message': 'Relatório gerado e enviado com sucesso!',
                'filename': filename,
                'period': f"{start_date} a {end_date}"
            }), 200
        else:
            return jsonify({'error': 'Erro ao enviar email. Tente novamente.'}), 500
            
    except ValueError as e:
        logger.warning(f"Validação falhou: {e}")
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logger.error(f"Erro ao processar relatório: {e}")
        return jsonify({'error': f'Erro interno do servidor: {str(e)}'}), 500

@reports_bp.route('/health', methods=['GET'])
def health_check():
    """Endpoint de health check."""
    return jsonify({
        'status': 'ok', 
        'message': 'Reports module is running',
        'timestamp': datetime.now().isoformat()
    }), 200
